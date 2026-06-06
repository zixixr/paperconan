"""Build a tiny xlsx fixture exercising several paperconan detectors.

Used by tests/test_smoke.py and for manual end-to-end testing:

    python tests/build_fixture.py tests/fixtures/tiny_paper
    paperconan tests/fixtures/tiny_paper

Patterns embedded:
  - `identical_column`        : col copy_of_mass ≡ col mass in every row
  - `arithmetic_progression`  : col ap_col is 2.5, 5.0, 7.5, … (non-integer step → high)
  - `cross_sheet_position_identical` : Sheet1 and Sheet2 share ≥ 15% bit-identical
                                       decimal values at the same (row, col)
  - `grim_inconsistent`       : Fig3_counts sheet — reported mean 3.45 at n=10 is
                                impossible for integer count data (GRIM)
"""
from __future__ import annotations

import os
import sys

import openpyxl


_ROWS = [
    # (mass, vol, copy_of_mass, ap_col)
    (1.2345, 2.5101, 1.2345, 2.5),
    (1.8923, 3.7842, 1.8923, 5.0),
    (2.5612, 4.9183, 2.5612, 7.5),
    (3.1456, 5.2734, 3.1456, 10.0),
    (3.8721, 6.5912, 3.8721, 12.5),
    (4.2389, 7.1148, 4.2389, 15.0),
    (4.9156, 8.3429, 4.9156, 17.5),
    (5.4823, 9.8765, 5.4823, 20.0),
    (6.1547, 10.5612, 6.1547, 22.5),
    (6.8392, 11.9234, 6.8392, 25.0),
]


def build(out_dir: str) -> str:
    """Write ED_Fig1.xlsx into out_dir; return the file path."""
    os.makedirs(out_dir, exist_ok=True)
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["mass(kg)", "vol(mL)", "copy_of_mass", "ap_col"])
    for r in _ROWS:
        ws1.append(list(r))

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["mass(kg)", "vol(mL)", "copy_of_mass", "ap_col"])
    # Same mass / copy_of_mass / vol as Sheet1 (triggers cross_sheet_position_identical),
    # except vol on row 8 (one byte-level tweak — common fabrication fingerprint),
    # and a different ap_col series so it isn't simply Sheet1 reused.
    for i, r in enumerate(_ROWS):
        mass, vol, copy_, _ = r
        if i == 7:
            vol = 9.0  # bit-tweak: differs at row 8, vol column
        ws2.append([mass, vol, copy_, 100.0 + i * 0.1])

    # Summary-statistics sheet with an integer-item keyword in the MEAN header and
    # one GRIM-impossible mean (3.45 is unreachable as an integer total / 10).
    ws3 = wb.create_sheet("Fig3_counts")
    ws3.append(["group", "cell count mean", "sd", "n"])
    ws3.append(["control", 3.40, 1.0, 10])   # consistent
    ws3.append(["treated", 3.45, 1.0, 10])   # GRIM-impossible
    ws3.append(["rescue", 3.30, 1.0, 10])    # consistent

    path = os.path.join(out_dir, "ED_Fig1.xlsx")
    wb.save(path)
    return path


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("usage: python tests/build_fixture.py <out_dir>", file=sys.stderr)
        return 2
    path = build(argv[1])
    print(f"wrote {path}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
