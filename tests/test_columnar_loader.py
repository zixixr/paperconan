import weakref
from zipfile import ZIP_DEFLATED, ZipFile

import numpy as np
import openpyxl
import pytest
from paperconan._audit import load_workbook_rows
from paperconan._sheet import Sheet


def _write_xlsx(path, rows, sheet_name="S1"):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = sheet_name
    for row in rows:
        ws.append(row)
    wb.save(path)


def _write_xlsx_with_exact_adjacent_wide_integers(path):
    _write_xlsx(path, [["a", "b"], [2**53, 2**53]])
    with ZipFile(path) as src:
        members = [(info, src.read(info.filename)) for info in src.infolist()]
    old = b'<c r="B2" t="n"><v>9007199254740992</v></c>'
    new = b'<c r="B2" t="n"><v>9007199254740993</v></c>'
    with ZipFile(path, "w", ZIP_DEFLATED) as dst:
        for info, data in members:
            if info.filename == "xl/worksheets/sheet1.xml":
                assert old in data
                data = data.replace(old, new, 1)
            dst.writestr(info, data)


def test_load_returns_sheets(tmp_path):
    p = tmp_path / "a.xlsx"
    _write_xlsx(p, [["H1", "H2"], [1, 2.5], [3, 4.0]])
    out = load_workbook_rows(str(p))
    s = out["S1"]
    assert isinstance(s, Sheet)
    assert s.cell(0, 0) == "H1"
    assert s.cell(1, 0) == 1 and isinstance(s.cell(1, 0), int)
    assert s.cell(1, 1) == 2.5
    assert s.nrows == 3 and s.ncols == 2


def test_oversized_sheet_is_none(tmp_path, monkeypatch):
    import paperconan._audit as A
    monkeypatch.setattr(A, "_MAX_CELLS", 5)
    p = tmp_path / "big.xlsx"
    _write_xlsx(p, [[i, i, i] for i in range(10)])   # 30 cells > 5
    out = load_workbook_rows(str(p))
    assert out["S1"] is None


def test_streaming_matches_from_rows(tmp_path):
    """The streamed Sheet must be byte-identical to Sheet.from_rows of the same
    data (numeric NaN-aware, text, ints, dims) — the parity guard for the rewrite."""
    rows = [["h", "k", "note"],
            [1, 2.5, "a"],
            [3, 4.0, None],
            [None, 0.001, "z"],
            [7, 8, "x"]]
    p = tmp_path / "p.xlsx"
    _write_xlsx(p, rows)
    streamed = load_workbook_rows(str(p))["S1"]
    # Build the reference the way scan_dir used to: read via openpyxl into rows, from_rows.
    wb = openpyxl.load_workbook(str(p), data_only=True, read_only=True)
    raw = [list(r) for r in wb["S1"].iter_rows(values_only=True)]
    wb.close()
    ref = Sheet.from_rows(raw)
    assert streamed.nrows == ref.nrows and streamed.ncols == ref.ncols
    assert np.array_equal(np.nan_to_num(streamed.numeric, nan=-123456.5),
                          np.nan_to_num(ref.numeric, nan=-123456.5))
    assert streamed._text == ref._text
    assert np.array_equal(streamed._ints, ref._ints)


def test_load_csv_returns_sheet(tmp_path):
    from paperconan._audit import load_csv_rows
    p = tmp_path / "d.csv"; p.write_text("a,b\n1,2.5\n3,x\n")
    s = load_csv_rows(str(p), delimiter=",")["d"]
    assert isinstance(s, Sheet)
    assert s.cell(0, 0) == "a"
    assert s.cell(1, 0) == 1 and isinstance(s.cell(1, 0), int)
    assert s.cell(1, 1) == 2.5
    assert s.cell(2, 1) == "x"


def test_csv_loader_streams_into_builder_without_sheet_from_rows(
    tmp_path, monkeypatch
):
    import paperconan._audit as audit

    path = tmp_path / "stream.csv"
    path.write_text("a,b\n1,2.5\n3,x\n", encoding="utf-8")

    def forbidden(_rows):
        raise AssertionError("CSV must not materialize normalized rows")

    monkeypatch.setattr(Sheet, "from_rows", forbidden)

    sheet = audit.load_csv_rows(str(path), ",")["stream"]

    assert isinstance(sheet, Sheet)
    assert sheet.cell(2, 1) == "x"


def test_csv_sparse_cell_limit_is_structured_and_exact(
    tmp_path, monkeypatch
):
    import paperconan._audit as audit

    path = tmp_path / "text.csv"
    path.write_text("a,b\nalpha,beta\n", encoding="utf-8")
    monkeypatch.setattr(audit, "_MAX_SPARSE_CELLS", 3, raising=False)
    monkeypatch.setattr(audit, "_MAX_SPARSE_BYTES", 100, raising=False)

    result = audit.load_table_result(str(path))

    assert result.sheets == {"text": None}
    assert [item.to_dict() for item in result.limitations] == [{
        "scope": "sheet",
        "reason": "sparse_cell_limit",
        "sheet": "text",
        "max_sparse_bytes": 100,
        "max_sparse_cells": 3,
        "observed_sparse_bytes": 11,
        "observed_sparse_cells": 4,
    }]


def test_csv_sparse_payload_limit_counts_exact_wide_integer_bytes(
    tmp_path, monkeypatch
):
    import paperconan._audit as audit

    path = tmp_path / "wide.csv"
    path.write_text("x\n9007199254740993\n", encoding="utf-8")
    monkeypatch.setattr(audit, "_MAX_SPARSE_CELLS", 10, raising=False)
    monkeypatch.setattr(audit, "_MAX_SPARSE_BYTES", 16, raising=False)

    result = audit.load_table_result(str(path))

    assert result.sheets == {"wide": None}
    assert [item.to_dict() for item in result.limitations] == [{
        "scope": "sheet",
        "reason": "sparse_payload_limit",
        "sheet": "wide",
        "max_sparse_bytes": 16,
        "max_sparse_cells": 10,
        "observed_sparse_bytes": 17,
        "observed_sparse_cells": 2,
    }]


def test_load_table_yields_sheets(tmp_path):
    from paperconan._audit import load_table
    p = tmp_path / "d.csv"; p.write_text("x\n1\n2\n3\n")
    out = load_table(str(p))
    assert all(v is None or isinstance(v, Sheet) for v in out.values())


def test_calamine_matches_openpyxl(tmp_path):
    import importlib.util, paperconan._audit as A
    if importlib.util.find_spec("python_calamine") is None:
        import pytest; pytest.skip("python-calamine not installed")
    p = tmp_path / "a.xlsx"
    _write_xlsx(p, [["H", "K"], [1, 2.5], [3, 4.0], [5, 6.25], [None, 7]])
    via_cal = A._load_workbook_calamine(str(p))
    via_op = A._load_workbook_openpyxl(str(p))
    for name in via_op:
        a, b = via_cal[name], via_op[name]
        assert (a is None) == (b is None)
        if a is not None:
            assert a.nrows == b.nrows and a.ncols == b.ncols
            assert np.array_equal(np.nan_to_num(a.numeric, nan=-1.5e9), np.nan_to_num(b.numeric, nan=-1.5e9))
            assert a._text == b._text
            assert np.array_equal(a._ints, b._ints)


def test_calamine_huge_bounding_box_is_none_not_oom(tmp_path, monkeypatch):
    import importlib.util
    if importlib.util.find_spec("python_calamine") is None:
        pytest.skip("no calamine")
    import paperconan._audit as A
    monkeypatch.setattr(A, "_MAX_CELLS", 1_000_000)
    p = tmp_path / "huge.xlsx"; wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"] = 1
    ws["C1000000"] = 2          # ~3M declared cells > 1M cap (within Excel's row limit)
    wb.save(str(p))

    # The oversized sheet must be rejected from its DECLARED dimensions before
    # its row stream starts.
    import python_calamine as pc
    orig = pc.CalamineSheet.iter_rows
    calls = {"n": 0}

    def spy(self, *a, **k):
        calls["n"] += 1
        return orig(self, *a, **k)

    def forbidden(*args, **kwargs):
        raise AssertionError("to_python must not be called")

    monkeypatch.setattr(pc.CalamineSheet, "iter_rows", spy)
    monkeypatch.setattr(pc.CalamineSheet, "to_python", forbidden)
    out = A._load_workbook_calamine(str(p))   # must NOT OOM; oversized -> None
    assert out["Sheet"] is None
    assert calls["n"] == 0                     # never started the huge row stream


def test_calamine_actual_size_guard_precedes_wide_ooxml_fallback(monkeypatch):
    import paperconan._audit as audit
    import python_calamine

    class StubSheet:
        height = 1
        width = 1

        def iter_rows(self):
            yield from [
                [float(2**53), 1.0],
                [2.0, 3.0],
                [4.0, 5.0],
            ]

        def to_python(self, skip_empty_area=False):
            raise AssertionError("to_python must not be called")

    class StubWorkbook:
        sheet_names = ["S1"]

        def get_sheet_by_name(self, name):
            assert name == "S1"
            return StubSheet()

    class StubCalamineWorkbook:
        @staticmethod
        def from_path(path):
            return StubWorkbook()

    fallback_calls = []

    def openpyxl_spy(path):
        fallback_calls.append(path)
        return {"fallback": Sheet.from_rows([])}

    monkeypatch.setattr(audit, "_MAX_CELLS", 5)
    monkeypatch.setattr(python_calamine, "CalamineWorkbook", StubCalamineWorkbook)
    monkeypatch.setattr(audit, "_load_workbook_openpyxl", openpyxl_spy)

    out = audit._load_workbook_calamine("oversized.xlsx")

    assert fallback_calls == []
    assert out["S1"] is None


def test_calamine_releases_reader_state_before_openpyxl_fallback(monkeypatch):
    import paperconan._audit as audit
    import python_calamine

    refs = {}

    class Probe:
        pass

    class ProbeRows:
        def __init__(self, values):
            self._values = iter(values)

        def __iter__(self):
            return self

        def __next__(self):
            return next(self._values)

    class StubSheet:
        height = 1
        width = 1

        def __init__(self, name, values):
            self.name = name
            self.values = values

        def iter_rows(self):
            rows = ProbeRows(self.values)
            refs[f"{self.name}_rows"] = weakref.ref(rows)
            return rows

        def to_python(self, skip_empty_area=False):
            raise AssertionError("to_python must not be called")

    class StubWorkbook:
        sheet_names = ["first", "wide"]

        def __init__(self):
            self.sheets = {
                "first": StubSheet("first", [[1.0]]),
                "wide": StubSheet("wide", [[float(2**53)]]),
            }
            refs["workbook"] = weakref.ref(self)
            for name, sheet in self.sheets.items():
                refs[f"{name}_sheet"] = weakref.ref(sheet)

        def get_sheet_by_name(self, name):
            return self.sheets[name]

    class StubCalamineWorkbook:
        @staticmethod
        def from_path(path):
            return StubWorkbook()

    def fill_spy(rows_iter, mr, mc, loaded):
        list(rows_iter)
        partial = Probe()
        refs["partial_output"] = weakref.ref(partial)
        return partial, mr * mc

    def openpyxl_spy(path):
        retained = sorted(name for name, ref in refs.items() if ref() is not None)
        assert retained == []
        return {"fallback": Sheet.from_rows([])}

    monkeypatch.setattr(python_calamine, "CalamineWorkbook", StubCalamineWorkbook)
    monkeypatch.setattr(audit, "_fill_sheet_from_rows", fill_spy)
    monkeypatch.setattr(audit, "_load_workbook_openpyxl", openpyxl_spy)

    out = audit._load_workbook_calamine("wide.xlsx")

    assert list(out) == ["fallback"]


def test_streaming_loader_preserves_adjacent_wide_integers(tmp_path):
    import paperconan._audit as audit

    p = tmp_path / "wide.xlsx"
    _write_xlsx_with_exact_adjacent_wide_integers(p)
    sheet = audit._load_workbook_openpyxl(str(p))["S1"]
    assert sheet.cell(1, 0) == 2**53
    assert sheet.cell(1, 1) == 2**53 + 1
    assert sheet.cell(1, 0) != sheet.cell(1, 1)


@pytest.mark.parametrize("mode", ["success", "row_iteration", "conversion"])
def test_openpyxl_loader_closes_workbook_exactly_once(
    monkeypatch, mode
):
    import paperconan._audit as audit

    expected = RuntimeError(f"{mode} failed")

    class StubSheet:
        max_row = 2
        max_column = 1

        def iter_rows(self, values_only=True):
            assert values_only is True
            yield [1]
            if mode == "row_iteration":
                raise expected
            yield [2]

    class StubWorkbook:
        sheetnames = ["S1"]

        def __init__(self):
            self.close_calls = 0

        def __getitem__(self, name):
            assert name == "S1"
            return StubSheet()

        def close(self):
            self.close_calls += 1

    workbook = StubWorkbook()
    monkeypatch.setattr(
        audit.openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: workbook,
    )
    if mode == "conversion":
        monkeypatch.setattr(
            audit,
            "_fill_sheet_from_rows",
            lambda *_args, **_kwargs: (_ for _ in ()).throw(expected),
        )

    if mode == "success":
        assert audit._load_workbook_openpyxl("input.xlsx")["S1"] is not None
    else:
        with pytest.raises(RuntimeError) as exc_info:
            audit._load_workbook_openpyxl("input.xlsx")
        assert exc_info.value is expected

    assert workbook.close_calls == 1


def test_openpyxl_loader_preserves_processing_error_if_close_fails(
    monkeypatch,
):
    import paperconan._audit as audit

    processing_error = RuntimeError("row iteration failed")

    class StubSheet:
        max_row = 1
        max_column = 1

        def iter_rows(self, values_only=True):
            assert values_only is True
            raise processing_error
            yield

    class StubWorkbook:
        sheetnames = ["S1"]

        def __getitem__(self, name):
            assert name == "S1"
            return StubSheet()

        def close(self):
            raise OSError("close failed")

    monkeypatch.setattr(
        audit.openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: StubWorkbook(),
    )

    with pytest.raises(RuntimeError) as exc_info:
        audit._load_workbook_openpyxl("input.xlsx")

    assert exc_info.value is processing_error


@pytest.mark.parametrize("suffix", [".xlsx", ".xlsm"])
def test_default_loader_falls_back_to_openpyxl_for_wide_ooxml_integers(
    tmp_path, monkeypatch, suffix
):
    import paperconan._audit as audit

    p = tmp_path / f"wide{suffix}"
    _write_xlsx_with_exact_adjacent_wide_integers(p)
    openpyxl_load = audit._load_workbook_openpyxl
    calls = []

    def spy(path):
        calls.append(path)
        return openpyxl_load(path)

    monkeypatch.setattr(audit, "_load_workbook_openpyxl", spy)

    sheet = audit.load_workbook_rows(str(p))["S1"]

    assert calls == [str(p)]
    assert sheet.cell(1, 0) == 2**53
    assert sheet.cell(1, 1) == 2**53 + 1
    assert sheet.cell(1, 0) != sheet.cell(1, 1)


@pytest.mark.parametrize("suffix", [".xls", ".xlsb"])
def test_legacy_workbooks_do_not_use_wide_integer_openpyxl_fallback(
    tmp_path, monkeypatch, suffix
):
    import paperconan._audit as audit
    import python_calamine

    class StubSheet:
        height = 1
        width = 1

        def iter_rows(self):
            yield [float(2**53)]

        def to_python(self, skip_empty_area=False):
            raise AssertionError("to_python must not be called")

    class StubWorkbook:
        sheet_names = ["S1"]

        def get_sheet_by_name(self, name):
            assert name == "S1"
            return StubSheet()

    class StubCalamineWorkbook:
        @staticmethod
        def from_path(path):
            return StubWorkbook()

    def forbidden(path):
        raise AssertionError("legacy workbook must stay on the Calamine path")

    monkeypatch.setattr(python_calamine, "CalamineWorkbook", StubCalamineWorkbook)
    monkeypatch.setattr(audit, "_load_workbook_openpyxl", forbidden)

    sheet = audit.load_workbook_rows(str(tmp_path / f"wide{suffix}"))["S1"]

    assert sheet.cell(0, 0) == 2**53


def test_ragged_csv_budget_uses_dense_geometry(tmp_path, monkeypatch):
    import paperconan._audit as audit

    monkeypatch.setattr(audit, "_MAX_CELLS", 12)
    path = tmp_path / "ragged.csv"
    path.write_text(
        "a\nb\nc\n" + ",".join(str(i) for i in range(5)) + "\n",
        encoding="utf-8",
    )
    assert audit.load_csv_rows(str(path), ",")["ragged"] is None


def test_second_workbook_sheet_is_rejected_before_allocation(tmp_path, monkeypatch):
    import paperconan._audit as audit

    monkeypatch.setattr(audit, "_MAX_CELLS", 10)
    fill_calls = []
    original_fill = audit._fill_sheet_from_rows

    def fill_spy(rows_iter, mr, mc, loaded):
        fill_calls.append((mr, mc, loaded))
        return original_fill(rows_iter, mr, mc, loaded)

    monkeypatch.setattr(audit, "_fill_sheet_from_rows", fill_spy)
    path = tmp_path / "two.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "one"
    ws1.append([1, 2, 3])
    ws1.append([4, 5, 6])
    ws2 = wb.create_sheet("two")
    ws2.append([1, 2, 3])
    ws2.append([4, 5, 6])
    wb.save(path)
    out = audit._load_workbook_openpyxl(str(path))
    assert out["one"] is not None
    assert out["two"] is None
    assert fill_calls == [(2, 3, 0)]


def test_calamine_streams_rows_without_to_python(tmp_path, monkeypatch):
    import python_calamine as pc
    import paperconan._audit as audit

    path = tmp_path / "a.xlsx"
    _write_xlsx(path, [["a", "b"], [1, 2], [3, 4]])

    def forbidden(*args, **kwargs):
        raise AssertionError("to_python must not be called")

    monkeypatch.setattr(pc.CalamineSheet, "to_python", forbidden)
    assert audit._load_workbook_calamine(str(path))["S1"] is not None


def test_calamine_second_sheet_rejected_before_iterator_starts(monkeypatch):
    import paperconan._audit as audit
    import python_calamine

    iterator_starts = []

    class StubSheet:
        height = 2
        width = 3

        def __init__(self, name):
            self.name = name

        def iter_rows(self):
            iterator_starts.append(self.name)
            yield [1.0, 2.0, 3.0]
            yield [4.0, 5.0, 6.0]

    class StubWorkbook:
        sheet_names = ["one", "two"]

        def __init__(self):
            self.sheets = {name: StubSheet(name) for name in self.sheet_names}

        def get_sheet_by_name(self, name):
            return self.sheets[name]

    class StubCalamineWorkbook:
        @staticmethod
        def from_path(path):
            return StubWorkbook()

    monkeypatch.setattr(audit, "_MAX_CELLS", 10)
    monkeypatch.setattr(python_calamine, "CalamineWorkbook", StubCalamineWorkbook)

    out = audit._load_workbook_calamine("two.xlsx")

    assert out["one"] is not None
    assert out["two"] is None
    assert iterator_starts == ["one"]


def test_ooxml_exception_releases_calamine_state_before_openpyxl_fallback(
    monkeypatch,
):
    import paperconan._audit as audit
    import python_calamine

    refs = {}

    class ReaderFailure(RuntimeError):
        pass

    class ProbeRows:
        def __iter__(self):
            return self

        def __next__(self):
            raise ReaderFailure("stream failed")

    class StubSheet:
        height = 1
        width = 1

        def __init__(self):
            refs["sheet"] = weakref.ref(self)

        def iter_rows(self):
            rows = ProbeRows()
            refs["iterator"] = weakref.ref(rows)
            return rows

    class StubWorkbook:
        sheet_names = ["S1"]

        def __init__(self):
            self.sheet = StubSheet()
            refs["workbook"] = weakref.ref(self)

        def get_sheet_by_name(self, name):
            assert name == "S1"
            return self.sheet

    class StubCalamineWorkbook:
        @staticmethod
        def from_path(path):
            return StubWorkbook()

    def openpyxl_spy(path):
        retained = sorted(name for name, ref in refs.items() if ref() is not None)
        assert retained == []
        return {"fallback": Sheet.from_rows([])}

    monkeypatch.setattr(python_calamine, "CalamineWorkbook", StubCalamineWorkbook)
    monkeypatch.setattr(audit, "_load_workbook_openpyxl", openpyxl_spy)

    out = audit.load_workbook_rows("broken.xlsx")

    assert list(out) == ["fallback"]


@pytest.mark.parametrize("suffix", [".xls", ".xlsb"])
def test_legacy_reader_exception_does_not_fallback_to_openpyxl(
    monkeypatch, suffix
):
    import paperconan._audit as audit
    import python_calamine

    class ReaderFailure(RuntimeError):
        pass

    class StubCalamineWorkbook:
        @staticmethod
        def from_path(path):
            raise ReaderFailure("legacy read failed")

    def forbidden(path):
        raise AssertionError("legacy reader errors must not use openpyxl")

    monkeypatch.setattr(python_calamine, "CalamineWorkbook", StubCalamineWorkbook)
    monkeypatch.setattr(audit, "_load_workbook_openpyxl", forbidden)

    with pytest.raises(ReaderFailure, match="legacy read failed"):
        audit.load_workbook_rows(f"broken{suffix}")


@pytest.mark.parametrize("suffix", [".xls", ".xlsb"])
def test_legacy_calamine_import_failure_does_not_fallback_to_openpyxl(
    monkeypatch, suffix
):
    import builtins
    import paperconan._audit as audit

    original_import = builtins.__import__

    def blocked_import(name, *args, **kwargs):
        if name == "python_calamine":
            raise ModuleNotFoundError("python_calamine unavailable")
        return original_import(name, *args, **kwargs)

    def forbidden(path):
        raise AssertionError("legacy import errors must not use openpyxl")

    monkeypatch.setattr(builtins, "__import__", blocked_import)
    monkeypatch.setattr(audit, "_load_workbook_openpyxl", forbidden)

    with pytest.raises(ModuleNotFoundError, match="python_calamine unavailable"):
        audit.load_workbook_rows(f"broken{suffix}")
