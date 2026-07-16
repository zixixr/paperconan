from pathlib import Path
import zipfile
from xml.etree import ElementTree as ET

import openpyxl
import pytest

import paperconan._audit as audit
from paperconan._audit import (
    _load_table_sheets,
    load_table,
    load_table_result,
    scan_dir,
)
import paperconan._input as input_module
from paperconan._input import inspect_ooxml_formula_cache
from paperconan._sheet import Sheet


_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_ROW_TAG = f"{{{_MAIN_NS}}}row"
_CELL_TAG = f"{{{_MAIN_NS}}}c"
_SHEET_DATA_TAG = f"{{{_MAIN_NS}}}sheetData"


def _write_formula_book(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stats"
    ws["A1"] = 2
    ws["A2"] = 3
    ws["A3"] = "=SUM(A1:A2)"
    wb.save(path)


def _rewrite_zip_member(path, member_name, transform):
    with zipfile.ZipFile(path) as src:
        members = [
            (info, src.read(info.filename))
            for info in src.infolist()
        ]
    with zipfile.ZipFile(path, "w") as dst:
        for info, data in members:
            if info.filename == member_name:
                data = transform(data)
            dst.writestr(info, data)


def _set_formula_cache(path, value):
    def transform(data):
        xml = ET.fromstring(data)
        cell = xml.find(f".//{{{_MAIN_NS}}}c[@r='A3']")
        assert cell is not None
        cached = cell.find(f"{{{_MAIN_NS}}}v")
        if cached is None:
            cached = ET.SubElement(cell, f"{{{_MAIN_NS}}}v")
        cached.text = value
        return ET.tostring(xml, encoding="utf-8", xml_declaration=True)

    _rewrite_zip_member(path, "xl/worksheets/sheet1.xml", transform)


def _set_first_worksheet_target(path, target):
    def transform(data):
        xml = ET.fromstring(data)
        relationship = xml.find(f"{{{_PKG_REL_NS}}}Relationship")
        assert relationship is not None
        relationship.attrib["Target"] = target
        return ET.tostring(xml, encoding="utf-8", xml_declaration=True)

    _rewrite_zip_member(path, "xl/_rels/workbook.xml.rels", transform)


@pytest.mark.parametrize("suffix", [".xlsx", ".xlsm"])
def test_formula_without_cached_value_is_reported(tmp_path, suffix):
    path = tmp_path / f"formula{suffix}"
    _write_formula_book(path)

    gaps = inspect_ooxml_formula_cache(str(path))

    assert gaps == {"Stats": {"count": 1, "cells": ["A3"]}}


def test_formula_gap_marks_scan_partial_without_counting_failures(tmp_path):
    data = tmp_path / "data"
    data.mkdir()
    path = data / "formula.xlsx"
    _write_formula_book(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Stats"]
    ws["B1"] = 11
    ws["B2"] = 12
    ws["B3"] = 13
    wb.save(path)
    wb.close()

    scan = scan_dir(str(data), str(tmp_path / "out"), write_html=False)

    assert scan["scan_status"] == "partial"
    assert scan["coverage"]["files_succeeded"] == 1
    assert scan["coverage"]["files_failed"] == 0
    assert scan["coverage"]["sheets_succeeded"] == 1
    assert scan["coverage"]["sheets_skipped"] == 0
    assert scan["coverage"]["limitations"] == [{
        "scope": "sheet",
        "reason": "formula_cache_missing",
        "file": "formula.xlsx",
        "sheet": "Stats",
        "count": 1,
        "cells": ["A3"],
    }]


def test_oversized_formula_sheet_keeps_structural_rejection_reason(
    tmp_path, monkeypatch
):
    data = tmp_path / "data"
    data.mkdir()
    path = data / "formula.xlsx"
    _write_formula_book(path)
    wb = openpyxl.load_workbook(path)
    wb["Stats"]["B3"] = 7
    wb.save(path)
    wb.close()
    monkeypatch.setattr(
        "paperconan._audit._MAX_CELLS", 5
    )

    scan = scan_dir(
        str(data), str(tmp_path / "out"), write_html=False
    )

    assert scan["coverage"]["sheets_skipped"] == 1
    assert [
        item["reason"]
        for item in scan["coverage"]["limitations"]
    ] == ["cell_limit"]
    assert scan["coverage"]["limitations"][0] == {
        "scope": "sheet",
        "reason": "cell_limit",
        "file": "formula.xlsx",
        "sheet": "Stats",
        "cells": 6,
        "max_cells": 5,
    }
    assert scan["scan_stats"]["sheets"][0]["oversized"] is True


def test_present_formula_cache_is_not_reported(tmp_path):
    path = tmp_path / "cached.xlsx"
    _write_formula_book(path)
    _set_formula_cache(path, "5")

    assert inspect_ooxml_formula_cache(str(path)) == {}


@pytest.mark.parametrize("value", [None, "", "   "])
def test_missing_or_empty_formula_cache_is_reported(tmp_path, value):
    path = tmp_path / "empty.xlsx"
    _write_formula_book(path)
    _set_formula_cache(path, value)

    assert inspect_ooxml_formula_cache(str(path)) == {
        "Stats": {"count": 1, "cells": ["A3"]}
    }


def test_zero_formula_cache_is_present(tmp_path):
    path = tmp_path / "zero.xlsx"
    _write_formula_book(path)
    _set_formula_cache(path, "0")

    assert inspect_ooxml_formula_cache(str(path)) == {}


def test_formula_gap_examples_are_bounded(tmp_path):
    path = tmp_path / "many.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stats"
    for row in range(1, 7):
        ws.cell(row, 1, f"={row}+1")
    wb.save(path)

    gaps = inspect_ooxml_formula_cache(str(path), max_examples=3)

    assert gaps == {
        "Stats": {
            "count": 6,
            "cells": ["A1", "A2", "A3"],
        }
    }


def test_formula_gap_examples_can_be_disabled(tmp_path):
    path = tmp_path / "many.xlsx"
    _write_formula_book(path)

    gaps = inspect_ooxml_formula_cache(str(path), max_examples=0)

    assert gaps == {"Stats": {"count": 1, "cells": []}}


def test_formula_cache_parser_detaches_processed_rows_and_cells(
    tmp_path, monkeypatch
):
    path = tmp_path / "streaming.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stats"
    for row in range(1, 51):
        ws.cell(row, 1, f"={row}+1")
    wb.save(path)

    real_iterparse = ET.iterparse
    detached_cells = []
    detached_rows = []

    def probing_iterparse(source, events):
        requested = set(events)
        current_row = None
        sheet_data = None
        for event, elem in real_iterparse(
            source, events=("start", "end")
        ):
            if event == "start":
                if elem.tag == _SHEET_DATA_TAG:
                    sheet_data = elem
                elif elem.tag == _ROW_TAG:
                    current_row = elem
            if event not in requested:
                continue
            yield event, elem
            if event == "end" and elem.tag == _CELL_TAG:
                assert current_row is not None
                detached_cells.append(elem not in current_row)
            elif event == "end" and elem.tag == _ROW_TAG:
                assert sheet_data is not None
                detached_rows.append(elem not in sheet_data)
                current_row = None

    monkeypatch.setattr(input_module.ET, "iterparse", probing_iterparse)

    gaps = inspect_ooxml_formula_cache(str(path))

    assert gaps == {
        "Stats": {
            "count": 50,
            "cells": [f"A{row}" for row in range(1, 21)],
        }
    }
    assert detached_cells == [True] * 50
    assert detached_rows == [True] * 50


def test_formula_cache_metadata_reads_are_bounded(tmp_path, monkeypatch):
    path = tmp_path / "bounded.xlsx"
    _write_formula_book(path)
    real_open = zipfile.ZipFile.open
    read_sizes = []
    guarded_members = {
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
    }

    class GuardedStream:
        def __init__(self, stream):
            self._stream = stream

        def read(self, size=-1):
            read_sizes.append(size)
            assert size >= 0
            return self._stream.read(size)

        def __enter__(self):
            self._stream.__enter__()
            return self

        def __exit__(self, *args):
            return self._stream.__exit__(*args)

        def __getattr__(self, name):
            return getattr(self._stream, name)

    def guarded_open(archive, member, *args, **kwargs):
        stream = real_open(archive, member, *args, **kwargs)
        name = (
            member.filename
            if isinstance(member, zipfile.ZipInfo)
            else member
        )
        if name in guarded_members:
            return GuardedStream(stream)
        return stream

    monkeypatch.setattr(zipfile.ZipFile, "open", guarded_open)

    assert inspect_ooxml_formula_cache(str(path)) == {
        "Stats": {"count": 1, "cells": ["A3"]}
    }
    assert read_sizes


def test_formula_cache_metadata_byte_limit_keeps_loaded_sheets(
    tmp_path, monkeypatch
):
    path = tmp_path / "bounded-metadata.xlsx"
    _write_formula_book(path)
    monkeypatch.setattr(
        input_module,
        "_OOXML_FORMULA_METADATA_BYTES",
        1,
        raising=False,
    )

    result = load_table_result(str(path))

    assert isinstance(result.sheets["Stats"], Sheet)
    assert [
        limitation.to_dict()
        for limitation in result.limitations
    ] == [{
        "scope": "file",
        "reason": "formula_metadata_byte_limit",
        "limit": 1,
        "member": "xl/workbook.xml",
    }]


def test_formula_cache_selected_sheet_limit_keeps_loaded_sheets(
    tmp_path, monkeypatch
):
    path = tmp_path / "many-sheets.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "First"
    wb.active["A1"] = "=1+1"
    wb.create_sheet("Second")["A1"] = "=2+2"
    wb.save(path)
    monkeypatch.setattr(
        input_module,
        "_OOXML_FORMULA_SHEET_LIMIT",
        1,
        raising=False,
    )
    monkeypatch.setattr(
        audit,
        "_try_load_workbook_calamine",
        lambda *_args, **_kwargs: audit._CALAMINE_READER_ERROR,
    )

    result = load_table_result(str(path))

    assert set(result.sheets) == {"First", "Second"}
    assert all(
        isinstance(sheet, Sheet)
        for sheet in result.sheets.values()
    )
    assert [
        limitation.to_dict()
        for limitation in result.limitations
    ] == [{
        "scope": "file",
        "reason": "formula_metadata_sheet_limit",
        "limit": 1,
        "selected_sheets": 2,
    }]


def test_formula_cache_path_resolution_retains_only_accepted_sheets(
    tmp_path,
):
    path = tmp_path / "selected-paths.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Accepted"
    wb.create_sheet("Rejected")
    wb.save(path)

    with zipfile.ZipFile(path) as archive:
        paths = input_module._worksheet_paths(
            archive,
            accepted_sheets={"Accepted"},
            max_sheets=1,
            metadata_byte_limit=1024 * 1024,
        )

    assert paths == [
        ("Accepted", "xl/worksheets/sheet1.xml")
    ]


def test_formula_cache_skips_unaccepted_worksheet(tmp_path, monkeypatch):
    path = tmp_path / "selected.xlsx"
    wb = openpyxl.Workbook()
    accepted = wb.active
    accepted.title = "Accepted"
    accepted["A1"] = "=1+1"
    rejected = wb.create_sheet("Rejected")
    rejected["A1"] = "=2+2"
    wb.save(path)

    real_open = zipfile.ZipFile.open

    def guarded_open(archive, member, *args, **kwargs):
        name = (
            member.filename
            if isinstance(member, zipfile.ZipInfo)
            else member
        )
        if name == "xl/worksheets/sheet2.xml":
            raise AssertionError("unaccepted worksheet was opened")
        return real_open(archive, member, *args, **kwargs)

    monkeypatch.setattr(zipfile.ZipFile, "open", guarded_open)

    assert inspect_ooxml_formula_cache(
        str(path),
        accepted_sheets={"Accepted"},
    ) == {
        "Accepted": {"count": 1, "cells": ["A1"]}
    }


def test_rejected_sheets_do_not_trigger_formula_inspection(
    tmp_path, monkeypatch
):
    path = tmp_path / "rejected.xlsx"
    path.write_bytes(b"placeholder")
    monkeypatch.setattr(
        audit,
        "_load_table_sheets",
        lambda _path, *, _limitations=None: {"Rejected": None},
    )

    def reject_inspection(*_args, **_kwargs):
        raise AssertionError("formula inspection should be skipped")

    monkeypatch.setattr(
        audit,
        "inspect_ooxml_formula_cache",
        reject_inspection,
    )

    result = load_table_result(str(path))

    assert result.sheets == {"Rejected": None}


def test_compatibility_loader_skips_formula_inspection(
    tmp_path, monkeypatch
):
    path = tmp_path / "compat.xlsx"
    path.write_bytes(b"placeholder")
    sheet = Sheet.from_rows([["value"], [1]])
    monkeypatch.setattr(
        audit,
        "_load_table_sheets",
        lambda _path, *, _limitations=None: {"Stats": sheet},
    )

    def reject_inspection(*_args, **_kwargs):
        raise AssertionError("compatibility loader inspected formulas")

    monkeypatch.setattr(
        audit,
        "inspect_ooxml_formula_cache",
        reject_inspection,
    )

    assert load_table(str(path)) == {"Stats": sheet}


def test_formula_gap_order_follows_workbook_and_cell_order(tmp_path):
    path = tmp_path / "ordered.xlsx"
    wb = openpyxl.Workbook()
    first = wb.active
    first.title = "First"
    first["B2"] = "=2+2"
    first["A1"] = "=1+1"
    second = wb.create_sheet("Second")
    second["C3"] = "=3+3"
    wb.save(path)

    gaps = inspect_ooxml_formula_cache(str(path))

    assert list(gaps) == ["First", "Second"]
    assert gaps["First"]["cells"] == ["A1", "B2"]
    assert gaps["Second"]["cells"] == ["C3"]


@pytest.mark.parametrize(
    "target",
    [
        "/xl/worksheets/sheet1.xml",
        r"worksheets\sheet1.xml",
        "worksheets/../worksheets/sheet1.xml",
    ],
)
def test_worksheet_relationship_targets_are_normalized(tmp_path, target):
    path = tmp_path / "normalized.xlsx"
    _write_formula_book(path)
    _set_first_worksheet_target(path, target)

    assert inspect_ooxml_formula_cache(str(path)) == {
        "Stats": {"count": 1, "cells": ["A3"]}
    }


def test_worksheet_relationship_cannot_leave_package(tmp_path):
    path = tmp_path / "traversal.xlsx"
    _write_formula_book(path)
    _set_first_worksheet_target(path, "../../outside.xml")

    with pytest.raises(ValueError, match="worksheet target leaves package"):
        inspect_ooxml_formula_cache(str(path))


def test_non_ooxml_input_has_no_formula_cache_gaps(tmp_path):
    path = tmp_path / "data.csv"
    path.write_text("a\n1\n", encoding="utf-8")

    assert inspect_ooxml_formula_cache(Path(path)) == {}


@pytest.mark.parametrize(
    ("cached", "expected_values", "expected_limitations"),
    [
        (
            False,
            [2, 3],
            [{
                "scope": "sheet",
                "reason": "formula_cache_missing",
                "sheet": "Stats",
                "cells": ["A3"],
                "count": 1,
            }],
        ),
        (True, [2, 3, 5], []),
    ],
    ids=["uncached-formula", "cached-formula"],
)
def test_ooxml_result_preserves_legacy_sheet_content(
    tmp_path, cached, expected_values, expected_limitations
):
    path = tmp_path / "formula.xlsx"
    _write_formula_book(path)
    if cached:
        _set_formula_cache(path, "5")

    baseline = _load_table_sheets(str(path))
    legacy = load_table(str(path))
    result = load_table_result(str(path))

    for sheets in (baseline, legacy, result.sheets):
        assert list(sheets) == ["Stats"]
        sheet = sheets["Stats"]
        assert isinstance(sheet, Sheet)
        assert (sheet.nrows, sheet.ncols) == (len(expected_values), 1)
        assert [
            sheet.cell(row, 0)
            for row in range(sheet.nrows)
        ] == expected_values
    assert [
        limitation.to_dict()
        for limitation in result.limitations
    ] == expected_limitations
