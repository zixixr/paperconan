"""Cell-count guard (OOM fix) + wide-block O(col^2) skip (disk/compute fix)."""
import csv

import numpy as np
import openpyxl

import paperconan._audit as A


def _make_xlsx(path, nrows, ncols):
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(nrows):
        ws.append([float(i * ncols + j) + 0.123 for j in range(ncols)])
    wb.save(path)


def test_oversized_sheet_skipped_and_recorded(tmp_path, monkeypatch):
    monkeypatch.setattr(A, "_MAX_CELLS", 50)            # tiny cap
    _make_xlsx(str(tmp_path / "big.xlsx"), 20, 10)      # 200 cells > 50
    scan = A.scan_dir(str(tmp_path), str(tmp_path / "out"), write_html=False)
    assert any("oversized sheet" in e.get("error", "") for e in scan["scan_errors"])
    assert any(s.get("oversized") for s in scan["scan_stats"]["sheets"])
    assert scan["n_blocks_with_findings"] == 0          # the skipped sheet produced nothing
    assert any(
        item["reason"] == "cell_limit"
        for item in scan["coverage"]["limitations"]
    )


def test_normal_sheet_under_cap_is_audited(tmp_path, monkeypatch):
    monkeypatch.setattr(A, "_MAX_CELLS", 2_000_000)
    _make_xlsx(str(tmp_path / "ok.xlsx"), 10, 6)         # 60 cells, fine
    scan = A.scan_dir(str(tmp_path), str(tmp_path / "out"), write_html=False)
    assert not any(s.get("oversized") for s in scan["scan_stats"]["sheets"])


def test_wide_block_skips_oncol2_detectors(tmp_path, monkeypatch):
    monkeypatch.setattr(A, "_MAX_BLOCK_COLS", 5)         # blocks wider than 5 cols skip relations
    monkeypatch.setattr(A, "_MAX_CELLS", 10_000_000)
    p = tmp_path / "wide.csv"
    with open(p, "w", newline="") as fh:
        w = csv.writer(fh)
        for i in range(8):
            row = [float(i) + 0.11 * j for j in range(12)]   # 12-col block (> 5)
            row[1] = row[0]                                   # col1 == col0 (identical_column bait)
            w.writerow(row)
    scan = A.scan_dir(str(tmp_path), str(tmp_path / "out2"), write_html=False)
    rels = [r for b in scan["relations_blocks"] for r in b.get("relations", [])]
    assert rels == []   # the O(col^2) relation detector was skipped on the wide block


def _track_dense_allocations(monkeypatch, limit, loaded):
    original_vstack = A.np.vstack
    original_hstack = A.np.hstack
    live = {}

    def shape_cells(shape):
        return int(np.prod(shape))

    def live_cells():
        return sum(record["size"] for record in live.values())

    def check_allocation(shape):
        requested = shape_cells(shape)
        assert requested <= limit - loaded
        assert loaded + live_cells() + requested <= limit

    class TrackedArray(np.ndarray):
        def __array_finalize__(self, source):
            self._allocation_record = getattr(
                source, "_allocation_record", None
            )
            self._allocation_owner = False

        def __del__(self):
            if getattr(self, "_allocation_owner", False):
                record = getattr(self, "_allocation_record", None)
                if record is not None:
                    live.pop(id(record), None)

        def resize(self, new_shape, refcheck=True):
            record = self._allocation_record
            requested = shape_cells(new_shape)
            assert requested <= limit - loaded
            assert (
                loaded + live_cells() - record["size"] + requested <= limit
            )
            result = super().resize(new_shape, refcheck=refcheck)
            record["size"] = requested
            return result

        def copy(self, order="C"):
            check_allocation(self.shape)
            result = new_tracked(
                self.shape,
                self.dtype,
                "F" if order == "F" else "C",
            )
            result[...] = self
            return result

    def new_tracked(shape, dtype, order):
        result = np.ndarray.__new__(
            TrackedArray,
            shape,
            dtype=dtype,
            order=order,
        )
        record = {"size": shape_cells(shape)}
        result._allocation_record = record
        result._allocation_owner = True
        live[id(record)] = record
        return result

    def tracked_full(shape, fill_value, dtype=None, order="C", **kwargs):
        check_allocation(shape)
        result = new_tracked(
            shape, float if dtype is None else dtype, order
        )
        result.fill(fill_value)
        return result

    def tracked_empty(shape, dtype=float, order="C", **kwargs):
        check_allocation(shape)
        return new_tracked(shape, dtype, order)

    def guarded_vstack(arrays, *args, **kwargs):
        rows = sum(array.shape[0] for array in arrays)
        cols = max((array.shape[1] for array in arrays), default=0)
        check_allocation((rows, cols))
        return original_vstack(arrays, *args, **kwargs)

    def guarded_hstack(arrays, *args, **kwargs):
        rows = max((array.shape[0] for array in arrays), default=0)
        cols = sum(array.shape[1] for array in arrays)
        check_allocation((rows, cols))
        return original_hstack(arrays, *args, **kwargs)

    monkeypatch.setattr(A.np, "full", tracked_full)
    monkeypatch.setattr(A.np, "empty", tracked_empty)
    monkeypatch.setattr(A.np, "vstack", guarded_vstack)
    monkeypatch.setattr(A.np, "hstack", guarded_hstack)


def _assert_compact_sheet(sheet, cells, shape):
    assert sheet is not None
    assert cells == np.prod(shape)
    assert sheet.numeric.shape == shape
    assert sheet.numeric.size == cells
    assert sheet.numeric.flags.owndata
    assert sheet.numeric.base is None


def test_column_growth_compacts_overdeclared_rows_before_allocation(monkeypatch):
    monkeypatch.setattr(A, "_MAX_CELLS", 10)
    _track_dense_allocations(monkeypatch, limit=10, loaded=4)

    sheet, cells = A._fill_sheet_from_rows(
        [[11], [21.5], [31, 32]], mr=6, mc=1, loaded=4
    )

    _assert_compact_sheet(sheet, cells, (3, 2))
    assert sheet.cell(0, 0) == 11
    assert isinstance(sheet.cell(0, 0), int)
    assert sheet.cell(0, 1) is None
    assert sheet.cell(1, 0) == 21.5
    assert isinstance(sheet.cell(1, 0), float)
    assert sheet.cell(1, 1) is None
    assert sheet.cell(2, 0) == 31
    assert sheet.cell(2, 1) == 32


def test_row_growth_compacts_overdeclared_columns_before_allocation(monkeypatch):
    monkeypatch.setattr(A, "_MAX_CELLS", 10)
    _track_dense_allocations(monkeypatch, limit=10, loaded=4)

    sheet, cells = A._fill_sheet_from_rows(
        [[11], [21], [31]], mr=2, mc=3, loaded=4
    )

    _assert_compact_sheet(sheet, cells, (3, 1))
    assert [sheet.cell(row, 0) for row in range(3)] == [11, 21, 31]
    assert all(isinstance(sheet.cell(row, 0), int) for row in range(3))


def test_overdeclared_sheet_returns_compact_owning_numeric_array(monkeypatch):
    monkeypatch.setattr(A, "_MAX_CELLS", 6)
    _track_dense_allocations(monkeypatch, limit=6, loaded=0)

    sheet, cells = A._fill_sheet_from_rows(
        [[11], [21.5]], mr=2, mc=3, loaded=0
    )

    _assert_compact_sheet(sheet, cells, (2, 1))
    assert sheet.cell(0, 0) == 11
    assert isinstance(sheet.cell(0, 0), int)
    assert sheet.cell(1, 0) == 21.5
    assert isinstance(sheet.cell(1, 0), float)
